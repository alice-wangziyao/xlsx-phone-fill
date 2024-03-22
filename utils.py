from openpyxl import load_workbook
import re
import os
from phone import Phone
from PyQt5.QtWidgets import QMessageBox

def Locationinformation(fp :str):

    if not os.path.exists(fp):
        QMessageBox.warning(None,"错误",f"不存在文件{fp}")
        return

    open_excel = load_workbook(fp, data_only=True)
    sheet = open_excel.sheetnames[0]
    worksheet = open_excel.active
    table = open_excel[sheet]
    columns = table.max_column
    phone =Phone()    
    worksheet.cell(row=1, column=6, value = "归属地")
    for i in range(1, columns + 1):
        cell_value = table.cell(row=1, column=i).value
        # 测试组人员bug数量从"bug维护人"这列取值
        if cell_value == '联系方式':
            # 获取某个单元格的坐标
            coordinate = table.cell(row=1, column=i).coordinate
            # 列以字母表示，
            letter = re.match("[A-Za-z]", coordinate)
            column = letter.group()
            for row in range(2, table.max_row + 1):
                cell = table[column + str(row)]
                phone_number = cell.value
                print(phone_number)
                
                # {'province':'上海'，"city':"上海’，"zip_code":'268880'，'area code':'621',"phone type':'电信"}
                info = phone.find(phone_number)
                province = info.get('province')
                city = info.get('city')
                phone_type = info.get('phone_type')
                print(info)
                worksheet.cell(row=row, column=6, value=f"{province} {city} {phone_type}")
    head, fname = os.path.split(fp)
    origin_fname = fname.split(".")[0]
    new_fp = f'{head}/new_{origin_fname}.xlsx'
    open_excel.save(new_fp)

    QMessageBox.warning(None,"提示",f"转换后的文件已保存至{new_fp}")
    

         

