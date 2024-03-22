from openpyxl import load_workbook
import re,sys,json

import requests
def Locationinformation():
    #excel = request.files.get('file')
    open_excel = load_workbook("", data_only=True)
    sheet = open_excel.sheetnames[0]
    worksheet = open_excel.active
    table = open_excel[sheet]
    columns = table.max_column
    for i in range(1, columns + 1):
        cell_value = table.cell(row=1, column=i).value
        # 测试组人员bug数量从"bug维护人"这列取值
        if cell_value == '联系方式':
            # 获取某个单元格的坐标
            coordinate = table.cell(row=1, column=i).coordinate
            # 列以字母表示，
            letter = re.match("[A-Za-z]", coordinate)
            column = letter.group()
            url = 'https://cx.shouji.360.cn/phonearea.php'
            for row in range(2, table.max_row + 1):
                cell = table[column + str(row)]
                phone_number = cell.value
                print(phone_number)
                params = {"number": phone_number}
                res = requests.get(url, params=params)
                res = res.text
                res = bytes(res, encoding="utf-8").decode("unicode_escape")
                province = json.loads(res)["data"]["province"]
                city = json.loads(res)["data"]["city"]
                sp = json.loads(res)["data"]["sp"]
                worksheet.cell(row=row, column=6, value=f"{province} {city} {sp}")
                sys.exit(1)
            open_excel.save('new.xlsx')
            return 'ok'

Locationinformation()